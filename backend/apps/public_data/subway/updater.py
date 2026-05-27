"""Subway public-data updater for the current capston schema."""

from __future__ import annotations

import hashlib
import json
import os
import re
import socket
import time as sleep_time
from dataclasses import dataclass
from datetime import time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from django.contrib.gis.geos import Point
from django.db import connection, transaction
from openpyxl import load_workbook

from apps.public_data.exceptions import RateLimitedError, is_rate_limited_text
from apps.public_data.regions.models import Adong, Ldong
from apps.public_data.state import load_state, record_dataset_result, save_state
from apps.public_data.subway.models import (
    NearestSubwayAdong,
    NearestSubwayLdong,
    SubwayCongestion,
    SubwayStation,
)


DATA_DIR = Path(__file__).resolve().parents[3] / "data"
LINE9_CONGESTION_PATH = DATA_DIR / "subway_line9_congestion.xlsx"
SEOUL_STATION_URL = "http://openapi.seoul.go.kr:8088/{key}/json/subwayStationMaster/{start}/{end}/"
SEOUL_CONGESTION_URL = "http://openapi.seoul.go.kr:8088/{key}/json/TbSeoulmetroStConInfo/{start}/{end}/"
PAGE_SIZE = 1000
TOP_K_NEAREST = 3


@dataclass(frozen=True)
class SubwayUpdateOptions:
    dry_run: bool = True
    force: bool = False
    limit: int | None = None
    request_interval_seconds: float = 0.2
    request_timeout_seconds: float = 40.0


def _require_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"{name} is required")
    return value


def _get(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return None


def _normalize_id(value: Any) -> str:
    text = str(value or "").strip()
    if text.isdigit():
        return str(int(text))
    return text


def _normalize_line(value: Any) -> str:
    text = str(value or "").strip()
    match = re.search(r"(\d+)", text)
    if match:
        return f"{int(match.group(1))}호선"
    return text


def _normalize_name(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def _file_meta(path: Path) -> dict[str, Any]:
    data = path.read_bytes()
    return {
        "file": str(path),
        "file_hash": hashlib.sha256(data).hexdigest(),
        "file_size": len(data),
    }


def _request_json(url: str, *, timeout: float) -> Any:
    req = Request(url, headers={"User-Agent": "capston-public-data-updater/0.1"})
    try:
        with urlopen(req, timeout=timeout) as res:
            return json.loads(res.read().decode("utf-8", errors="replace"))
    except HTTPError as exc:
        if exc.code == 429:
            raise RateLimitedError("Seoul API rate limited: HTTP 429") from exc
        raise RuntimeError(f"Seoul API request failed: HTTP {exc.code}") from exc
    except (URLError, TimeoutError, socket.timeout) as exc:
        raise RuntimeError(f"Seoul API request failed: {type(exc).__name__}") from exc
    except json.JSONDecodeError as exc:
        raise RuntimeError("Seoul API returned malformed JSON") from exc


def _seoul_rows(service: str, url_template: str, api_key: str, options: SubwayUpdateOptions):
    start = 1
    fetched = 0
    while True:
        end = start + PAGE_SIZE - 1
        payload = _request_json(
            url_template.format(key=api_key, start=start, end=end),
            timeout=options.request_timeout_seconds,
        )
        body = payload.get(service) or {}
        result = body.get("RESULT") or {}
        if str(result.get("CODE", "")).startswith("ERROR"):
            if is_rate_limited_text(result):
                raise RateLimitedError(f"Seoul API rate limited for {service}: {result}")
            raise RuntimeError(f"Seoul API error for {service}: {result}")
        rows = body.get("row") or []
        for row in rows:
            if options.limit is not None and fetched >= options.limit:
                return
            fetched += 1
            yield row
        total = int(body.get("list_total_count") or 0)
        if not rows or start + len(rows) > total:
            return
        start += len(rows)
        if options.request_interval_seconds:
            sleep_time.sleep(options.request_interval_seconds)


def _find_region(point: Point) -> tuple[Ldong | None, Adong | None]:
    return (
        Ldong.objects.filter(boundary__covers=point).order_by("area_m2").first(),
        Adong.objects.filter(boundary__covers=point).order_by("area_m2").first(),
    )


def _build_station_records(options: SubwayUpdateOptions) -> dict[str, Any]:
    api_key = _require_env("SEOUL_API_KEY")
    records: list[dict[str, Any]] = []
    source_ids: set[str] = set()
    checked = skipped = 0
    skip_reasons = {"bad_coord": 0, "missing_required": 0}

    for row in _seoul_rows("subwayStationMaster", SEOUL_STATION_URL, api_key, options):
        checked += 1
        station_id = _normalize_id(_get(row, "STATN_ID", "STATION_CD", "FR_CODE", "BLDN_ID"))
        name = str(_get(row, "STATN_NM", "STATION_NM", "SBWY_STNS_NM", "BLDN_NM") or "").strip()
        line = str(_get(row, "ROUTE", "LINE_NUM", "LINE", "SBWY_ROUT_LN_NM") or "").strip()
        try:
            lon = float(_get(row, "CRDNT_X", "XPOINT_WGS", "LON", "LOT", "longitude"))
            lat = float(_get(row, "CRDNT_Y", "YPOINT_WGS", "LAT", "latitude"))
        except (TypeError, ValueError):
            skipped += 1
            skip_reasons["bad_coord"] += 1
            continue
        if not (station_id and name and line):
            skipped += 1
            skip_reasons["missing_required"] += 1
            continue
        point = Point(lon, lat, srid=4326)
        ldong, adong = _find_region(point)
        source_ids.add(station_id)
        records.append(
            {
                "id": station_id,
                "defaults": {
                    "name": name[:100],
                    "line": line[:20],
                    "location": point,
                    "ldong_id": ldong.ldong_code if ldong else None,
                    "adong_id": adong.adong_code if adong else None,
                },
                "region_found": bool(ldong and adong),
            }
        )

    return {
        "status": "success" if options.limit is None else "partial",
        "completed": options.limit is None,
        "checked": checked,
        "records": records,
        "source_ids": source_ids,
        "skipped": skipped,
        "skip_reasons": skip_reasons,
        "region_not_found": sum(1 for record in records if not record["region_found"]),
    }


def update_stations(options: SubwayUpdateOptions) -> dict[str, Any]:
    built = _build_station_records(options)
    created = updated = deleted_missing = 0
    loaded = len(built["records"]) if options.dry_run else 0

    if not options.dry_run and built["completed"]:
        with transaction.atomic():
            for record in built["records"]:
                _, was_created = SubwayStation.objects.update_or_create(
                    id=record["id"],
                    defaults=record["defaults"],
                )
                loaded += 1
                created += int(was_created)
                updated += int(not was_created)
            missing_qs = SubwayStation.objects.exclude(id__in=built["source_ids"])
            deleted_missing = missing_qs.count()
            missing_qs.delete()

    return {
        "status": built["status"],
        "completed": built["completed"],
        "checked": built["checked"],
        "loaded": loaded,
        "created": created,
        "updated": updated,
        "deleted_missing": deleted_missing,
        "skipped": built["skipped"],
        "skip_reasons": built["skip_reasons"],
        "region_not_found": built["region_not_found"],
        "dry_run": options.dry_run,
    }


def _parse_time_label(value: Any) -> time | None:
    text = str(value or "").strip()
    match = re.search(r"(\d{1,2})\s*(?::|시)\s*(\d{2})?", text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2) or 0)
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return time(hour, minute)


def _decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace("%", "").strip())
    except InvalidOperation:
        return None


def _station_lookup() -> tuple[dict[str, str], dict[tuple[str, str], str], dict[str, str]]:
    by_id: dict[str, str] = {}
    by_name_line: dict[tuple[str, str], str] = {}
    by_line9_name: dict[str, str] = {}
    for station in SubwayStation.objects.only("id", "name", "line").iterator():
        by_id[str(station.id)] = str(station.id)
        name_key = _normalize_name(station.name)
        line_key = _normalize_line(station.line)
        by_name_line[(name_key, line_key)] = str(station.id)
        if "9" in station.line:
            by_line9_name[name_key] = str(station.id)
    return by_id, by_name_line, by_line9_name


def _resolve_station_id(
    row: dict[str, Any],
    *,
    by_id: dict[str, str],
    by_name_line: dict[tuple[str, str], str],
) -> str | None:
    station_id = _normalize_id(_get(row, "STATN_ID", "STATION_CD", "STN_ID", "역번호", "역ID"))
    if station_id and station_id in by_id:
        return by_id[station_id]
    name = _normalize_name(_get(row, "STATN_NM", "STATION_NM", "SBWY_STNS_NM", "역명"))
    line = _normalize_line(_get(row, "LINE", "LINE_NUM", "ROUTE", "호선"))
    if name and line:
        return by_name_line.get((name, line))
    return None


def _parse_congestion_row(
    row: dict[str, Any],
    *,
    by_id: dict[str, str],
    by_name_line: dict[tuple[str, str], str],
) -> tuple[list[SubwayCongestion], str | None]:
    station_id = _resolve_station_id(row, by_id=by_id, by_name_line=by_name_line)
    if not station_id:
        return [], "station_not_found"

    day_type = str(_get(row, "DAY_TYPE", "DOW", "WKND_SE", "요일구분") or "").strip()
    direction = str(_get(row, "DIRECTION", "DIR", "UPDN_LINE", "상하구분") or "").strip()
    express_yn = str(_get(row, "EXPRESS_YN", "EXPRESS", "TRAIN_SE", "급행여부") or "일반").strip()
    if not (day_type and direction):
        return [], "missing_required"

    direct_time = _parse_time_label(_get(row, "HH", "TIME", "TM", "TIME_SLOT"))
    direct_value = _decimal(_get(row, "CONGESTION", "CGST", "CROWDED", "RCPTN_RATE", "혼잡도"))
    if direct_time and direct_value is not None:
        return [
            SubwayCongestion(
                station_id=station_id,
                day_type=day_type[:20],
                direction=direction[:20],
                express_yn=express_yn[:10],
                time=direct_time,
                congestion=direct_value,
            )
        ], None

    records: list[SubwayCongestion] = []
    for key, value in row.items():
        slot = _parse_time_label(key)
        congestion = _decimal(value)
        if slot is None or congestion is None:
            continue
        records.append(
            SubwayCongestion(
                station_id=station_id,
                day_type=day_type[:20],
                direction=direction[:20],
                express_yn=express_yn[:10],
                time=slot,
                congestion=congestion,
            )
        )
    if not records:
        return [], "missing_congestion"
    return records, None


def _build_api_congestion(options: SubwayUpdateOptions) -> dict[str, Any]:
    api_key = _require_env("SEOUL_API_KEY")
    by_id, by_name_line, _by_line9_name = _station_lookup()
    checked = skipped = loaded = 0
    skip_reasons = {"station_not_found": 0, "missing_required": 0, "missing_congestion": 0}
    records: list[SubwayCongestion] = []

    for row in _seoul_rows("TbSeoulmetroStConInfo", SEOUL_CONGESTION_URL, api_key, options):
        checked += 1
        parsed, reason = _parse_congestion_row(row, by_id=by_id, by_name_line=by_name_line)
        if reason:
            skipped += 1
            skip_reasons[reason] += 1
            continue
        loaded += len(parsed)
        records.extend(parsed)

    return {
        "status": "success" if options.limit is None else "partial",
        "completed": options.limit is None,
        "checked": checked,
        "loaded": loaded,
        "records": records,
        "skipped": skipped,
        "skip_reasons": skip_reasons,
    }


def _sheet_parts(sheet_name: str) -> tuple[str, str, str]:
    direction = "상선" if "상선" in sheet_name else "하선"
    express_yn = "급행" if "급행" in sheet_name else "일반"
    day_type = "평일" if "평일" in sheet_name else "휴일"
    return day_type, direction, express_yn


def _build_line9_congestion(options: SubwayUpdateOptions) -> dict[str, Any]:
    if not LINE9_CONGESTION_PATH.exists():
        raise RuntimeError(f"missing file: {LINE9_CONGESTION_PATH}")

    _by_id, _by_name_line, by_line9_name = _station_lookup()
    wb = load_workbook(LINE9_CONGESTION_PATH, read_only=True, data_only=True)
    checked = skipped = loaded = 0
    skip_reasons = {"station_not_found": 0, "bad_value": 0}
    records: list[SubwayCongestion] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        day_type, direction, express_yn = _sheet_parts(sheet_name)
        header = rows[1]
        time_cols = [(idx, _parse_time_label(label)) for idx, label in enumerate(header[1:], start=1)]
        for row in rows[2:]:
            station_name = _normalize_name(row[0])
            station_id = by_line9_name.get(station_name)
            if not station_id:
                skipped += 1
                skip_reasons["station_not_found"] += 1
                continue
            for idx, slot in time_cols:
                if slot is None or idx >= len(row) or row[idx] in (None, ""):
                    continue
                checked += 1
                congestion = _decimal(row[idx])
                if congestion is None:
                    skipped += 1
                    skip_reasons["bad_value"] += 1
                    continue
                records.append(
                    SubwayCongestion(
                        station_id=station_id,
                        day_type=day_type,
                        direction=direction,
                        express_yn=express_yn,
                        time=slot,
                        congestion=congestion,
                    )
                )
                loaded += 1
                if options.limit is not None and checked >= options.limit:
                    return {
                        "status": "partial",
                        "completed": False,
                        "checked": checked,
                        "loaded": loaded,
                        "records": records,
                        "skipped": skipped,
                        "skip_reasons": skip_reasons,
                        "file": _file_meta(LINE9_CONGESTION_PATH),
                    }

    return {
        "status": "success",
        "completed": True,
        "checked": checked,
        "loaded": loaded,
        "records": records,
        "skipped": skipped,
        "skip_reasons": skip_reasons,
        "file": _file_meta(LINE9_CONGESTION_PATH),
    }


def _dedupe_congestion(records: list[SubwayCongestion]) -> tuple[list[SubwayCongestion], int]:
    by_key: dict[tuple[str, str, str, str, time], SubwayCongestion] = {}
    for record in records:
        key = (
            str(record.station_id),
            record.day_type,
            record.direction,
            record.express_yn,
            record.time,
        )
        by_key[key] = record
    return list(by_key.values()), len(records) - len(by_key)


def update_congestion(options: SubwayUpdateOptions) -> dict[str, Any]:
    api = _build_api_congestion(options)
    line9 = _build_line9_congestion(options)
    completed = bool(api["completed"] and line9["completed"])
    records, duplicate_rows = _dedupe_congestion(api["records"] + line9["records"])
    replaced = 0

    if completed and not options.dry_run:
        with transaction.atomic():
            replaced = SubwayCongestion.objects.count()
            SubwayCongestion.objects.all().delete()
            SubwayCongestion.objects.bulk_create(
                records,
                batch_size=5000,
            )
    elif options.dry_run:
        replaced = SubwayCongestion.objects.count()

    return {
        "status": "success" if completed else "partial",
        "completed": completed,
        "loaded": len(records) if (completed or options.dry_run) else 0,
        "duplicate_rows": duplicate_rows,
        "replaced_existing": replaced,
        "dry_run": options.dry_run,
        "congestion_api": {key: value for key, value in api.items() if key != "records"},
        "line9_file": {key: value for key, value in line9.items() if key != "records"},
    }


def _nearest_rows(region_table: str, pk_column: str) -> list[tuple[str, int, str, float]]:
    sql = f"""
        SELECT r.{pk_column}, nearest.name, nearest.distance_m
        FROM {region_table} r
        CROSS JOIN LATERAL (
            SELECT s.name, ST_Distance(s.location::geography, r.location::geography) AS distance_m
            FROM subway_station s
            WHERE r.location IS NOT NULL
            ORDER BY s.location <-> r.location
            LIMIT %s
        ) nearest
        ORDER BY r.{pk_column}, nearest.distance_m
    """
    with connection.cursor() as cur:
        cur.execute(sql, [TOP_K_NEAREST])
        raw = cur.fetchall()

    out: list[tuple[str, int, str, float]] = []
    current_code = None
    rank = 0
    for code, station_name, distance_m in raw:
        if code != current_code:
            current_code = code
            rank = 1
        else:
            rank += 1
        out.append((str(code), rank, str(station_name), float(distance_m)))
    return out


def update_nearest(options: SubwayUpdateOptions) -> dict[str, Any]:
    if not SubwayStation.objects.exists():
        raise RuntimeError("subway stations must be loaded before nearest cache")

    adong_rows = _nearest_rows("adong", "adong_code")
    ldong_rows = _nearest_rows("ldong", "ldong_code")
    deleted_adong = deleted_ldong = 0

    if not options.dry_run:
        with transaction.atomic():
            deleted_adong = NearestSubwayAdong.objects.count()
            deleted_ldong = NearestSubwayLdong.objects.count()
            NearestSubwayAdong.objects.all().delete()
            NearestSubwayLdong.objects.all().delete()
            NearestSubwayAdong.objects.bulk_create(
                [
                    NearestSubwayAdong(
                        adong_id=code,
                        rank=rank,
                        station_name=station_name[:100],
                        distance_m=distance_m,
                    )
                    for code, rank, station_name, distance_m in adong_rows
                ],
                batch_size=3000,
            )
            NearestSubwayLdong.objects.bulk_create(
                [
                    NearestSubwayLdong(
                        ldong_id=code,
                        rank=rank,
                        station_name=station_name[:100],
                        distance_m=distance_m,
                    )
                    for code, rank, station_name, distance_m in ldong_rows
                ],
                batch_size=3000,
            )

    return {
        "status": "success",
        "completed": True,
        "loaded": {
            "nearest_subway_adong": len(adong_rows),
            "nearest_subway_ldong": len(ldong_rows),
        },
        "deleted_existing": {
            "nearest_subway_adong": deleted_adong,
            "nearest_subway_ldong": deleted_ldong,
        },
        "dry_run": options.dry_run,
    }


def update_subway(options: SubwayUpdateOptions) -> dict[str, Any]:
    if options.dry_run:
        stations = update_stations(options)
        congestion = update_congestion(options)
        nearest = update_nearest(options) if congestion["completed"] else None
    else:
        with transaction.atomic():
            stations = update_stations(options)
            if not stations["completed"]:
                return {
                    "status": "partial",
                    "completed": False,
                    "loaded": stations["loaded"],
                    "dry_run": options.dry_run,
                    "stations": stations,
                    "congestion": None,
                    "nearest": None,
                }
            congestion = update_congestion(options)
            nearest = update_nearest(options) if congestion["completed"] else None
    completed = bool(stations["completed"] and congestion["completed"] and nearest and nearest["completed"])
    return {
        "status": "success" if completed else "partial",
        "completed": completed,
        "loaded": stations["loaded"] + congestion["loaded"] + (sum(nearest["loaded"].values()) if nearest else 0),
        "dry_run": options.dry_run,
        "stations": stations,
        "congestion": congestion,
        "nearest": nearest,
    }


def update(options: SubwayUpdateOptions) -> dict[str, Any]:
    result: dict[str, Any] = {
        "dataset": "subway",
        "dry_run": options.dry_run,
        "subway": None,
    }
    try:
        result["subway"] = update_subway(options)
    except Exception as exc:
        result["error"] = {
            "type": type(exc).__name__,
            "message": str(exc),
        }
        if not options.dry_run:
            record_dataset_result("subway", result)
        raise

    if not options.dry_run:
        record_dataset_result("subway", result)
        if result["subway"] and result["subway"].get("completed"):
            state = load_state()
            subway_state = state.setdefault("datasets", {}).setdefault("subway", {})
            subway_state["stations_snapshot"] = {
                "last_success_date": None,
                "checked": result["subway"]["stations"]["checked"],
                "loaded": result["subway"]["stations"]["loaded"],
                "deleted_missing": result["subway"]["stations"]["deleted_missing"],
                "skip_reasons": result["subway"]["stations"]["skip_reasons"],
            }
            subway_state["congestion_snapshot"] = {
                "last_success_date": None,
                "loaded": result["subway"]["congestion"]["loaded"],
                "replaced_existing": result["subway"]["congestion"]["replaced_existing"],
                "line9_file": result["subway"]["congestion"]["line9_file"]["file"],
            }
            subway_state["nearest_snapshot"] = result["subway"]["nearest"]["loaded"]
            save_state(state)
    return result
