"""Store public-data updater for the current capston schema."""

from __future__ import annotations

import hashlib
import json
import os
import socket
import time as sleep_time
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from django.contrib.gis.geos import Point
from django.db import transaction
from openpyxl import load_workbook

from apps.public_data.exceptions import RateLimitedError, is_rate_limited_code, is_rate_limited_text
from apps.public_data.regions.models import Adong, Ldong
from apps.public_data.state import dataset_state, load_state, record_dataset_result, save_state
from apps.public_data.store.models import BusinessCategory, KsciCategory, Store


DATA_DIR = Path(__file__).resolve().parents[3] / "data"
BUSINESS_CATEGORY_PATH = DATA_DIR / "store_business_category.xlsx"
KSCI_CATEGORY_PATH = DATA_DIR / "KSIC_10th.xlsx"
STORE_API_URL = "https://apis.data.go.kr/B553077/api/open/sdsc2/storeListInDong"
PAGE_SIZE = 1000


@dataclass(frozen=True)
class StoresUpdateOptions:
    dry_run: bool = True
    force: bool = False
    limit: int | None = None
    request_interval_seconds: float = 0.2
    request_timeout_seconds: float = 40.0


def _require_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"{name} is required")
    return value


def _file_meta(path: Path) -> dict[str, Any]:
    data = path.read_bytes()
    return {
        "file": str(path),
        "file_hash": hashlib.sha256(data).hexdigest(),
        "file_size": len(data),
    }


def _catalog_snapshot() -> dict[str, Any]:
    business = _file_meta(BUSINESS_CATEGORY_PATH)
    ksci = _file_meta(KSCI_CATEGORY_PATH)
    digest = hashlib.sha256()
    for item in (business, ksci):
        digest.update(item["file"].encode("utf-8"))
        digest.update(b"\0")
        digest.update(item["file_hash"].encode("ascii"))
        digest.update(b"\0")
    return {
        "file_hash": digest.hexdigest(),
        "files": {
            "store_business_category.xlsx": business,
            "KSIC_10th.xlsx": ksci,
        },
    }


def _business_rows() -> list[dict[str, str]]:
    wb = load_workbook(BUSINESS_CATEGORY_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        main_code, main_name, middle_code, middle_name, sub_code, sub_name = row[:6]
        if not sub_code:
            continue
        rows.append(
            {
                "main_code": str(main_code).strip(),
                "main_name": str(main_name or "").strip(),
                "middle_code": str(middle_code).strip(),
                "middle_name": str(middle_name or "").strip(),
                "sub_code": str(sub_code).strip(),
                "sub_name": str(sub_name or "").strip(),
            }
        )
    if len(rows) < 200:
        raise RuntimeError(f"refusing incomplete business category file: {len(rows)} rows")
    return rows


def _ksci_rows() -> list[dict[str, str]]:
    wb = load_workbook(KSCI_CATEGORY_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    current = {"main_code": "", "main_name": "", "middle_name": "", "sub_name": "", "class_name": ""}
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        main_code, main_name, _middle_code, middle_name, _sub_code, sub_name, _class_code, class_name, detail_code, detail_name = row[:10]
        if main_code:
            current["main_code"] = str(main_code).strip()
            current["main_name"] = str(main_name or "").strip()
        if middle_name:
            current["middle_name"] = str(middle_name).strip()
        if sub_name:
            current["sub_name"] = str(sub_name).strip()
        if class_name:
            current["class_name"] = str(class_name).strip()
        if not detail_code:
            continue
        rows.append(
            {
                "ksci_code": f"{current['main_code']}{str(detail_code).strip()}",
                "subcategory_name": str(detail_name or "").strip(),
                "class_name": current["class_name"],
                "subclass_name": current["sub_name"],
                "middle_category_name": current["middle_name"],
                "main_category_name": current["main_name"],
            }
        )
    if len(rows) < 1000:
        raise RuntimeError(f"refusing incomplete KSIC file: {len(rows)} rows")
    return rows


def _load_catalog(options: StoresUpdateOptions, snapshot: dict[str, Any]) -> dict[str, Any]:
    business_rows = _business_rows()
    ksci_rows = _ksci_rows()
    previous_hash = dataset_state("stores").get("catalog_snapshot", {}).get("file_hash")
    skipped_write = bool(previous_hash == snapshot["file_hash"] and not options.force and not options.dry_run)
    loaded_business = loaded_ksci = 0

    if not skipped_write and not options.dry_run:
        with transaction.atomic():
            for row in business_rows:
                BusinessCategory.objects.update_or_create(
                    subcategory_code=row["sub_code"],
                    defaults={
                        "subcategory_name": row["sub_name"],
                        "middle_category_code": row["middle_code"],
                        "middle_category_name": row["middle_name"],
                        "main_category_code": row["main_code"],
                        "main_category_name": row["main_name"],
                    },
                )
                loaded_business += 1

            for row in ksci_rows:
                KsciCategory.objects.update_or_create(
                    ksci_code=row["ksci_code"],
                    defaults={
                        "subcategory_name": row["subcategory_name"],
                        "class_name": row["class_name"],
                        "subclass_name": row["subclass_name"],
                        "middle_category_name": row["middle_category_name"],
                        "main_category_name": row["main_category_name"],
                    },
                )
                loaded_ksci += 1

    return {
        "status": "success",
        "completed": True,
        "checked": {"business_category": len(business_rows), "ksci_category": len(ksci_rows)},
        "loaded": {"business_category": loaded_business, "ksci_category": loaded_ksci},
        "skipped_write": skipped_write,
        "files": snapshot,
    }


def _request_json(params: dict[str, str], options: StoresUpdateOptions) -> dict[str, Any]:
    if options.request_interval_seconds:
        sleep_time.sleep(options.request_interval_seconds)
    req = Request(
        f"{STORE_API_URL}?{urlencode(params, safe='%')}",
        headers={"User-Agent": "capston-public-data-updater/0.1"},
    )
    try:
        with urlopen(req, timeout=options.request_timeout_seconds) as res:
            return json.loads(res.read().decode("utf-8", errors="replace"))
    except HTTPError as exc:
        if exc.code == 429:
            raise RateLimitedError("store API rate limited: HTTP 429") from exc
        raise RuntimeError(f"store API request failed: HTTP {exc.code}") from exc
    except (URLError, TimeoutError, socket.timeout) as exc:
        raise RuntimeError(f"store API request failed: {type(exc).__name__}") from exc
    except json.JSONDecodeError as exc:
        raise RuntimeError("store API returned malformed JSON") from exc


def _response_items(payload: dict[str, Any]) -> tuple[int, list[dict[str, Any]]]:
    header = payload.get("header") or {}
    result_code = str(header.get("resultCode") or "").strip()
    if result_code and result_code != "00":
        if is_rate_limited_code(result_code) or is_rate_limited_text(header):
            raise RateLimitedError(f"store API rate limited resultCode={result_code} resultMsg={header.get('resultMsg')}")
        raise RuntimeError(f"store API error resultCode={result_code} resultMsg={header.get('resultMsg')}")
    body = payload.get("body") or {}
    total = int(body.get("totalCount") or 0)
    items = body.get("items") or []
    if isinstance(items, dict):
        items = items.get("item") or []
    if isinstance(items, dict):
        items = [items]
    return total, items if isinstance(items, list) else []


def _fetch_gu_rows(api_key: str, gu_code: str, options: StoresUpdateOptions) -> list[dict[str, Any]]:
    page = 1
    collected: list[dict[str, Any]] = []
    while True:
        payload = _request_json(
            {
                "serviceKey": api_key,
                "divId": "signguCd",
                "key": gu_code,
                "pageNo": str(page),
                "numOfRows": str(PAGE_SIZE),
                "type": "json",
            },
            options,
        )
        total, rows = _response_items(payload)
        collected.extend(rows)
        if not rows or not total or page * PAGE_SIZE >= total:
            return collected
        page += 1


def _normalize_adong_code(value: Any) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if len(text) == 8:
        return f"{text}00"
    return text


def _point(row: dict[str, Any]) -> Point | None:
    lon = row.get("lon") or row.get("lonVal") or row.get("x")
    lat = row.get("lat") or row.get("latVal") or row.get("y")
    try:
        return Point(float(lon), float(lat), srid=4326)
    except (TypeError, ValueError):
        return None


def _find_region(point: Point, ldong_id: str | None, adong_id: str | None) -> tuple[Ldong | None, Adong | None]:
    ldong = Ldong.objects.filter(ldong_code=ldong_id).first() if ldong_id else None
    adong = Adong.objects.filter(adong_code=adong_id).first() if adong_id else None
    if not ldong:
        ldong = Ldong.objects.filter(boundary__covers=point).order_by("area_m2").first()
    if not adong:
        adong = Adong.objects.filter(boundary__covers=point).order_by("area_m2").first()
    return ldong, adong


def _build_store_record(
    row: dict[str, Any],
    *,
    category_ids: set[str],
    ksci_ids: set[str],
) -> tuple[dict[str, Any] | None, str | None, str | None]:
    store_id = str(row.get("bizesId") or "").strip()
    if not store_id:
        return None, None, "missing_id"
    point = _point(row)
    if not point:
        return None, store_id, "missing_location"
    category_id = str(row.get("indsSclsCd") or "").strip()
    if category_id not in category_ids:
        return None, store_id, "unknown_category"
    ldong_id = str(row.get("ldongCd") or "").strip() or None
    adong_id = _normalize_adong_code(row.get("adongCd"))
    ldong, adong = _find_region(point, ldong_id, adong_id)
    if not (ldong and adong):
        return None, store_id, "region_not_found"
    ksci_id = str(row.get("ksicCd") or "").strip() or None
    if ksci_id and ksci_id not in ksci_ids:
        ksci_id = None

    return (
        {
            "id": store_id[:50],
            "defaults": {
                "name": str(row.get("bizesNm") or "")[:100],
                "branch_name": str(row.get("brchNm") or "")[:100] or None,
                "address": str(row.get("rdnmAdr") or row.get("lnoAdr") or "")[:255],
                "location": point,
                "category_id": category_id,
                "ksci_id": ksci_id,
                "ldong_id": ldong.ldong_code,
                "adong_id": adong.adong_code,
            },
        },
        store_id[:50],
        None,
    )


def _fetch_and_build_stores(options: StoresUpdateOptions) -> dict[str, Any]:
    api_key = _require_env("PUBLIC_DATA_API_KEY")
    gu_codes = list(Ldong.objects.values_list("gu_id", flat=True).distinct().order_by("gu_id"))
    category_ids = set(BusinessCategory.objects.values_list("subcategory_code", flat=True))
    ksci_ids = set(KsciCategory.objects.values_list("ksci_code", flat=True))
    if not gu_codes:
        raise RuntimeError("regions must be loaded before updating stores")
    if not category_ids:
        raise RuntimeError("business categories must be loaded before updating stores")

    checked = skipped = 0
    source_ids: set[str] = set()
    records: list[dict[str, Any]] = []
    skip_reasons = {
        "missing_id": 0,
        "missing_location": 0,
        "unknown_category": 0,
        "region_not_found": 0,
    }
    completed = False

    for gu_code in gu_codes:
        rows = _fetch_gu_rows(api_key, gu_code, options)
        for row in rows:
            if options.limit is not None and checked >= options.limit:
                return {
                    "status": "partial",
                    "completed": False,
                    "checked": checked,
                    "records": records,
                    "source_ids": source_ids,
                    "skipped": skipped,
                    "skip_reasons": skip_reasons,
                    "gu_count": len(gu_codes),
                }
            checked += 1
            record, source_id, reason = _build_store_record(row, category_ids=category_ids, ksci_ids=ksci_ids)
            if source_id:
                source_ids.add(source_id)
            if not record:
                skipped += 1
                if reason:
                    skip_reasons[reason] += 1
                continue
            records.append(record)
    completed = True
    return {
        "status": "success",
        "completed": completed,
        "checked": checked,
        "records": records,
        "source_ids": source_ids,
        "skipped": skipped,
        "skip_reasons": skip_reasons,
        "gu_count": len(gu_codes),
    }


def update_stores_data(options: StoresUpdateOptions) -> dict[str, Any]:
    built = _fetch_and_build_stores(options)
    loaded = created = updated = deleted_missing = 0
    if not options.dry_run and built["completed"]:
        with transaction.atomic():
            for record in built["records"]:
                _, was_created = Store.objects.update_or_create(
                    id=record["id"],
                    defaults=record["defaults"],
                )
                loaded += 1
                created += int(was_created)
                updated += int(not was_created)
            if options.limit is None:
                missing_qs = Store.objects.exclude(id__in=built["source_ids"])
                deleted_missing = missing_qs.count()
                missing_qs.delete()
    elif not built["completed"]:
        loaded = 0
    else:
        loaded = len(built["records"])

    return {
        "status": "success" if built["completed"] else "partial",
        "completed": built["completed"],
        "checked": built["checked"],
        "loaded": loaded,
        "created": created,
        "updated": updated,
        "skipped": built["skipped"],
        "deleted_missing": deleted_missing,
        "dry_run": options.dry_run,
        "gu_count": built["gu_count"],
        "skip_reasons": built["skip_reasons"],
    }


def update_stores(options: StoresUpdateOptions) -> dict[str, Any]:
    catalog_snapshot = _catalog_snapshot()
    catalog = _load_catalog(options, catalog_snapshot)
    stores = update_stores_data(options)
    completed = bool(catalog["completed"] and stores["completed"])
    catalog_loaded = sum(catalog["loaded"].values())
    return {
        "status": "success" if completed else "partial",
        "completed": completed,
        "loaded": catalog_loaded + stores["loaded"],
        "dry_run": options.dry_run,
        "catalog": catalog,
        "stores": stores,
    }


def update(options: StoresUpdateOptions) -> dict[str, Any]:
    result: dict[str, Any] = {
        "dataset": "stores",
        "dry_run": options.dry_run,
        "stores": None,
    }
    try:
        result["stores"] = update_stores(options)
    except Exception as exc:
        result["error"] = {
            "type": type(exc).__name__,
            "message": str(exc),
        }
        if not options.dry_run:
            record_dataset_result("stores", result)
        raise

    if not options.dry_run:
        record_dataset_result("stores", result)
        if result["stores"] and result["stores"].get("completed"):
            state = load_state()
            stores_state = state.setdefault("datasets", {}).setdefault("stores", {})
            stores_state["catalog_snapshot"] = result["stores"]["catalog"]["files"] | {
                "loaded": result["stores"]["catalog"]["loaded"],
            }
            stores_state["stores_snapshot"] = {
                "last_success_date": None,
                "checked": result["stores"]["stores"]["checked"],
                "loaded": result["stores"]["stores"]["loaded"],
                "deleted_missing": result["stores"]["stores"]["deleted_missing"],
                "skip_reasons": result["stores"]["stores"]["skip_reasons"],
            }
            save_state(state)
    return result
